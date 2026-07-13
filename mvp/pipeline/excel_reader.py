import os
import re
import csv as _csv_stdlib
import json
import time
import urllib.parse
import openpyxl
from google import genai
from google.genai import types

_SYSTEM = (
    "You are a data mapping assistant for a background verification company.\n"
    "You receive the first few rows of an Excel file.\n\n"
    "Your job:\n"
    "1. Identify which row index (0-based) is the actual header row — "
    "the row containing column labels, not data or section titles.\n"
    "2. For each column in that header row, map it to one of these standard fields:\n"
    "   - ars_number    : the ARS case reference for this candidate (e.g. 'ARS no', 'case_ars_no', 'ARS Number')\n"
    "   - check_id      : the individual check or case ID (e.g. 'Case Check Id', 'Check No', 'CHK ID'). "
    "One candidate (ARS) may have multiple check rows, each with a different check_id.\n"
    "   - check_type    : the type or name of this background verification check "
    "(e.g. 'Employment', 'Education', 'Address', 'Criminal', 'Reference'). "
    "Typically labelled 'Check Type', 'Verification Type', 'Service Name', 'Component', or similar.\n"
    "   - candidate_name: the personal name of the candidate being verified — "
    "typically labelled 'Candidate Name', 'Full Name', 'First Name', or 'Last Name'. "
    "If the name is split (e.g. 'First Name' and 'Last Name'), map BOTH to candidate_name. "
    "Do NOT map employer or company name columns (e.g. 'Emp Name', 'Employer Name', 'Company') "
    "to candidate_name — those belong to employment history, not the candidate's identity.\n"
    "   - emp_id        : the candidate's employee or staff ID assigned by their employer\n"
    "   - unknown       : anything else — education details, employment history, dates, remarks\n\n"
    "Return only valid JSON. No explanation, no markdown, no text outside the JSON."
)

_SYSTEM_MULTI = (
    "You are a data mapping assistant for a background verification company.\n"
    "You receive sample rows (first ~10) from MULTIPLE Excel/CSV files that were attached "
    "to the same bulk-insufficiency email. The files may describe the same set of candidates "
    "from different angles (e.g. one has ARS numbers, another has Employee IDs for the same "
    "people), or may be unrelated — you are not told which.\n\n"
    "Your job, for EACH file independently:\n"
    "1. Identify which row index (0-based) is the actual header row — "
    "the row containing column labels, not data or section titles.\n"
    "2. For each column in that header row, map it to one of these standard fields:\n"
    "   - ars_number    : the ARS case reference for this candidate\n"
    "   - check_id      : the individual check or case ID. "
    "One candidate (ARS) may have multiple check rows, each with a different check_id.\n"
    "   - check_type    : the type or name of this background verification check "
    "(e.g. 'Employment', 'Education', 'Address', 'Criminal', 'Reference')\n"
    "   - candidate_name: the candidate's personal name — map First Name AND Last Name "
    "both here if the name is split. Never map employer/company name columns here.\n"
    "   - emp_id        : the candidate's employee or staff ID assigned by their employer\n"
    "   - unknown       : anything else\n\n"
    "3. ACROSS the files: if the standard fields above are not enough to tell which row in one "
    "file corresponds to which row in another file for the SAME candidate (e.g. neither file's "
    "ARS/EmpID/Name columns overlap, but you notice a different shared column such as a custom "
    "'Application Ref No' whose sample values clearly match across files), report that column pair "
    "as extra_link_field. Only report it if the sample values genuinely look like the same "
    "reference across files — do not guess. If the standard fields already look sufficient to "
    "link candidates across files, or no shared column is evident, return null for extra_link_field.\n\n"
    "Return only valid JSON. No explanation, no markdown, no text outside the JSON."
)


def _strip_fences(text):
    text = text.strip()
    if text.startswith('```'):
        parts = text.split('```')
        text = parts[1] if len(parts) > 1 else text
        if text.startswith('json'):
            text = text[4:]
    return text.strip()


_MODELS      = ['gemini-2.5-flash', 'gemini-2.5-flash-lite']
_RETRY_WAITS = [0, 5, 15, 30]

# Rows sent to the AI for the multi-file column-mapping / link-detection call.
# Kept small on purpose — this call only needs enough sample data to infer
# headers and a possible join key, not the full sheet.
_PEEK_SAMPLE_ROWS = 10


def _is_transient(e):
    s = str(e)
    return any(x in s for x in ('503', '429', 'UNAVAILABLE', 'RESOURCE_EXHAUSTED'))


def _llm_map_columns(peek_rows):
    """
    Send the first few Excel rows to Gemini. Returns:
      { "header_row_index": <int>, "column_map": { "<header text>": "<standard_field>" } }
    Retries on transient errors with the same schedule as grouping_agent.
    Falls back to gemini-2.5-flash-lite if the primary model is exhausted.
    Raises on second JSON parse failure or full retry exhaustion.
    """
    api_key = os.environ.get('GEMINI_API_KEY', '').strip()
    if not api_key:
        raise Exception(
            'GEMINI_API_KEY is not configured. '
            'AI column mapping requires the API key to be set in .env'
        )

    client = genai.Client(api_key=api_key)

    rows_text = json.dumps(
        [[str(c) if c is not None else '' for c in row] for row in peek_rows],
        ensure_ascii=False,
        indent=2,
    )

    base_prompt = (
        f"Here are the first rows of an Excel file:\n{rows_text}\n\n"
        "Return JSON in exactly this structure:\n"
        "{\n"
        '  "header_row_index": <0-based integer>,\n'
        '  "column_map": {\n'
        '    "<exact column header text as it appears in the row>": '
        '"<ars_number|check_id|candidate_name|emp_id|unknown>"\n'
        '  }\n'
        "}"
    )
    retry_prompt = (
        base_prompt + '\n\nCRITICAL: Your previous response was not valid JSON. '
        'Return only the JSON object — no markdown, no explanation.'
    )

    last_error = None
    json_fail_count = 0

    for model in _MODELS:
        for attempt, wait in enumerate(_RETRY_WAITS):
            if wait:
                time.sleep(wait)
            prompt = retry_prompt if json_fail_count > 0 else base_prompt
            try:
                response = client.models.generate_content(
                    model=model,
                    contents=prompt,
                    config=types.GenerateContentConfig(system_instruction=_SYSTEM),
                )
                return json.loads(_strip_fences(response.text))
            except json.JSONDecodeError as e:
                last_error = e
                json_fail_count += 1
                if json_fail_count >= 2:
                    raise Exception(
                        f'AI column mapping returned invalid JSON twice: {e}. '
                        'Check the Excel file — it may have an unusual structure.'
                    )
            except Exception as e:
                last_error = e
                if _is_transient(e):
                    print(f'[ExcelReader] {model} attempt {attempt + 1} — transient error: {e}')
                    if attempt == len(_RETRY_WAITS) - 1:
                        print(f'[ExcelReader] All retries for {model} exhausted, trying next model…')
                else:
                    print(f'[ExcelReader] {model} attempt {attempt + 1} — error: {e}')
                    break  # non-transient — skip remaining retries for this model
        else:
            continue
        break

    raise Exception(
        f'AI column mapping failed after all retries: {last_error}. '
        'Check the Excel file — it may have an unusual structure.'
    )


def _llm_map_columns_multi(peeks_by_filename):
    """
    Send sample rows (already capped by the caller) from MULTIPLE Excel/CSV files to
    Gemini in a single call. Returns:
      {
        "files": { "<filename>": {"header_row_index": <int>, "column_map": {...}}, ... },
        "extra_link_field": { "<filename>": "<exact header text in that file>", ... } | None
      }
    Same retry/model-fallback schedule as _llm_map_columns.
    """
    api_key = os.environ.get('GEMINI_API_KEY', '').strip()
    if not api_key:
        raise Exception(
            'GEMINI_API_KEY is not configured. '
            'AI column mapping requires the API key to be set in .env'
        )

    client = genai.Client(api_key=api_key)

    files_block = {
        fname: [[str(c) if c is not None else '' for c in row] for row in rows]
        for fname, rows in peeks_by_filename.items()
    }
    files_text = json.dumps(files_block, ensure_ascii=False, indent=2)

    base_prompt = (
        f"Here are sample rows from each attached file:\n{files_text}\n\n"
        "Return JSON in exactly this structure:\n"
        "{\n"
        '  "files": {\n'
        '    "<filename exactly as given above>": {\n'
        '      "header_row_index": <0-based integer>,\n'
        '      "column_map": {\n'
        '        "<exact column header text>": '
        '"<ars_number|check_id|check_type|candidate_name|emp_id|unknown>"\n'
        '      }\n'
        '    }\n'
        '  },\n'
        '  "extra_link_field": {\n'
        '    "<filename>": "<exact column header text in that file>"\n'
        '  }\n'
        "}\n"
        "(extra_link_field must include one entry per file that shares the join key, "
        "or be JSON null if no such shared column was found)"
    )
    retry_prompt = (
        base_prompt + '\n\nCRITICAL: Your previous response was not valid JSON. '
        'Return only the JSON object — no markdown, no explanation.'
    )

    last_error = None
    json_fail_count = 0

    for model in _MODELS:
        for attempt, wait in enumerate(_RETRY_WAITS):
            if wait:
                time.sleep(wait)
            prompt = retry_prompt if json_fail_count > 0 else base_prompt
            try:
                response = client.models.generate_content(
                    model=model,
                    contents=prompt,
                    config=types.GenerateContentConfig(system_instruction=_SYSTEM_MULTI),
                )
                result = json.loads(_strip_fences(response.text))
                result.setdefault('files', {})
                result.setdefault('extra_link_field', None)
                return result
            except json.JSONDecodeError as e:
                last_error = e
                json_fail_count += 1
                if json_fail_count >= 2:
                    raise Exception(
                        f'AI column mapping (multi-file) returned invalid JSON twice: {e}. '
                        'Check the Excel files — they may have an unusual structure.'
                    )
            except Exception as e:
                last_error = e
                if _is_transient(e):
                    print(f'[ExcelReader] {model} attempt {attempt + 1} (multi-file) — transient error: {e}')
                    if attempt == len(_RETRY_WAITS) - 1:
                        print(f'[ExcelReader] All retries for {model} exhausted, trying next model…')
                else:
                    print(f'[ExcelReader] {model} attempt {attempt + 1} (multi-file) — error: {e}')
                    break
        else:
            continue
        break

    raise Exception(
        f'AI column mapping (multi-file) failed after all retries: {last_error}. '
        'Check the Excel files — they may have an unusual structure.'
    )


def _find_data_sheet(wb):
    """
    Scan all sheets, peek first 10 rows from each, and return
    (worksheet, peek_rows) for the sheet with the most non-empty cells —
    the most likely candidate data sheet.
    Each iter_rows() call on a ReadOnlyWorksheet opens a fresh stream,
    so multiple calls are safe.
    """
    best_ws    = wb.active
    best_peek  = []
    best_score = -1

    for name in wb.sheetnames:
        ws   = wb[name]
        peek = []
        for row in ws.iter_rows(values_only=True):
            peek.append(row)
            if len(peek) >= 20:
                break
        score = sum(
            1 for row in peek
            for c in row
            if c is not None and str(c).strip()
        )
        if score > best_score:
            best_score = score
            best_ws    = ws
            best_peek  = peek

    return best_ws, best_peek


def parse_excel(excel_path):
    """
    Parse a candidate spreadsheet regardless of format.
    Supported: .xlsx (openpyxl), .xls (xlrd), .xlsb (pyxlsb), .ods (odfpy), .csv (stdlib csv).
    All formats share the same AI column-mapping and row-streaming logic.
    """
    ext = os.path.splitext(excel_path)[1].lower()
    if ext == '.csv':
        return _parse_csv(excel_path)
    if ext == '.xls':
        return _parse_xls(excel_path)
    if ext == '.xlsb':
        return _parse_xlsb(excel_path)
    if ext == '.ods':
        return _parse_ods(excel_path)
    # Default: .xlsx (and any unrecognised extension — openpyxl will raise a clear error)
    return _parse_xlsx(excel_path)


def parse_excel_multi(paths):
    """
    Parse multiple Excel/CSV attachments and cross-reference them into ONE candidate
    row list. Each file is still column-mapped by AI — one combined call across all
    files, using only the first _PEEK_SAMPLE_ROWS rows per file — then every file's
    full rows are streamed with the same per-format logic as parse_excel(). Rows that
    describe the same real candidate across files are linked with a deterministic
    cascading exact-ID match (ars > check_id > emp_id > name > AI-detected
    extra_link_field) and identity fields are backfilled across the match. check_id /
    check_type / raw_fields are never touched by the merge — they stay per-row, so the
    existing per-ARS multi-check grouping in app.py still applies once rows share an ars.

    Returns the same row shape as parse_excel(): a flat list of records with
    ars/check_id/check_type/name/emp_id/folder_key/folder_key_type/raw_fields,
    plus optional _merge_conflict / _merge_basis notes for the manifest.
    """
    peeked    = {}   # filename -> (peek, remaining_iter, cleanup)
    filenames = []   # preserves attachment upload order

    try:
        for path in paths:
            fname = os.path.basename(path)
            filenames.append(fname)
            peeked[fname] = _peek_file(path)

        sample_peeks   = {fname: peeked[fname][0][:_PEEK_SAMPLE_ROWS] for fname in filenames}
        mapping_result = _llm_map_columns_multi(sample_peeks)

        file_mappings    = mapping_result.get('files', {})
        extra_link_field = mapping_result.get('extra_link_field')

        all_rows = []
        for fname in filenames:
            mapping = file_mappings.get(fname)
            if not mapping:
                raise Exception(
                    f'AI column mapping (multi-file) did not return a mapping for "{fname}". '
                    f'Files mapped: {list(file_mappings.keys())}'
                )
            peek, remaining_iter, _cleanup = peeked[fname]
            rows = _stream_rows(peek, remaining_iter, mapping, require_identity=False)
            link_header = extra_link_field.get(fname) if extra_link_field else None
            for row in rows:
                row['_source_file'] = fname
                if link_header:
                    row['_link_value'] = row.get('raw_fields', {}).get(link_header, '')
                all_rows.append(row)
    finally:
        for _peek, _remaining_iter, cleanup in peeked.values():
            cleanup()

    _cross_link_rows(all_rows, filenames)
    return all_rows


# ── cross-file candidate matching ────────────────────────────────────────────

def _norm_name_for_match(s):
    return re.sub(r'[\s\-_.]', '', (s or '').lower())


def _pairwise_match(a, b):
    """
    Deterministic cascading exact match between two candidate rows from different
    files: ars -> check_id -> emp_id -> normalised name -> AI-detected link value.
    Returns True on the first field where both sides are non-blank and equal.
    A mismatch on a higher-priority field means "not the same candidate" even if a
    lower-priority field would coincidentally match, so this does not fall through
    once both sides have a non-blank value for a field.
    """
    fields = [
        lambda r: (r.get('ars')      or '').strip().lower(),
        lambda r: (r.get('check_id') or '').strip().lower(),
        lambda r: (r.get('emp_id')   or '').strip().lower(),
        lambda r: _norm_name_for_match(r.get('name')),
        lambda r: (r.get('_link_value') or '').strip().lower(),
    ]
    for extract in fields:
        va, vb = extract(a), extract(b)
        if va and vb:
            return va == vb
    return False


class _UnionFind:
    def __init__(self, n):
        self.parent = list(range(n))

    def find(self, x):
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, x, y):
        rx, ry = self.find(x), self.find(y)
        if rx != ry:
            self.parent[ry] = rx


def _cross_link_rows(all_rows, filenames):
    """
    Union rows from DIFFERENT source files that _pairwise_match() considers the same
    candidate, then backfill ars/emp_id/name across each linked group using
    first-non-blank-wins (in attachment upload order). Mutates all_rows in place.
    """
    n = len(all_rows)
    if n < 2:
        return

    uf = _UnionFind(n)
    for i in range(n):
        for j in range(i + 1, n):
            if all_rows[i].get('_source_file') == all_rows[j].get('_source_file'):
                continue  # same-file grouping is handled downstream in app.py, not here
            if _pairwise_match(all_rows[i], all_rows[j]):
                uf.union(i, j)

    file_order = {fname: idx for idx, fname in enumerate(filenames)}
    components = {}
    for i in range(n):
        components.setdefault(uf.find(i), []).append(i)

    for indices in components.values():
        if len(indices) < 2:
            continue
        indices.sort(key=lambda i: file_order.get(all_rows[i].get('_source_file'), 0))

        # Determine match basis BEFORE backfilling — backfill propagates ID fields
        # across the whole group, which would otherwise make every group look
        # "linked via ID" even when the original match was name-only.
        linked_via_id = False
        for a in range(len(indices)):
            for b in range(a + 1, len(indices)):
                ra, rb = all_rows[indices[a]], all_rows[indices[b]]
                for key in ('ars', 'check_id', 'emp_id', '_link_value'):
                    va = (ra.get(key) or '').strip().lower()
                    vb = (rb.get(key) or '').strip().lower()
                    if va and vb and va == vb:
                        linked_via_id = True
        if not linked_via_id:
            for i in indices:
                all_rows[i]['_merge_basis'] = 'name_only'

        for field in ('ars', 'emp_id', 'name'):
            values = [(i, (all_rows[i].get(field) or '').strip()) for i in indices]
            non_blank = [v for _, v in values if v]
            if not non_blank:
                continue
            winner = non_blank[0]
            if len(set(non_blank)) > 1:
                sources = [
                    f'{v} ({all_rows[i].get("_source_file")})'
                    for i, v in values if v
                ]
                conflict_note = f'{field} mismatch across files: ' + ' vs '.join(sources)
                for i in indices:
                    all_rows[i].setdefault('_merge_conflict', [])
                    all_rows[i]['_merge_conflict'].append(conflict_note)
            # Propagate the winning value onto every row in the group — not just
            # blanks — so all rows describing the same candidate stay consistent.
            for i in indices:
                all_rows[i][field] = winner

        # Recompute folder_key/folder_key_type now that ars/emp_id/name may have
        # been backfilled from the other file(s) in this group.
        for i in indices:
            row = all_rows[i]
            if row.get('ars'):
                row['folder_key'], row['folder_key_type'] = row['ars'], 'ars'
            elif row.get('check_id'):
                row['folder_key'], row['folder_key_type'] = row['check_id'], 'check_id'
            elif row.get('emp_id'):
                row['folder_key'], row['folder_key_type'] = row['emp_id'], 'emp_id'
            elif row.get('name'):
                row['folder_key'], row['folder_key_type'] = row['name'], 'name'
            # A row that started with no identifier of its own (require_identity=False
            # let it through) may now have one via backfill — clear the stale flag so
            # it isn't mistaken for still-unresolved.
            if row.get('folder_key') and row.get('_flag') == 'no_identifier':
                del row['_flag']


# ── peek extraction (format-specific, no AI call) ────────────────────────────
# Each _peek_* returns (peek_rows, remaining_row_iterator, cleanup) without
# invoking the AI mapper — used both by the single-file _parse_* wrappers below
# and directly by parse_excel_multi() to gather samples from every file before
# making one combined AI call.

def _peek_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        return _peek_csv(path)
    if ext == '.xls':
        return _peek_xls(path)
    if ext == '.xlsb':
        return _peek_xlsb(path)
    if ext == '.ods':
        return _peek_ods(path)
    return _peek_xlsx(path)


def _peek_csv(path):
    """Peek a CSV file using stdlib csv — no AI needed for reading, but AI still maps columns."""
    rows = []
    with open(path, newline='', encoding='utf-8-sig', errors='replace') as f:
        reader = _csv_stdlib.reader(f)
        for row in reader:
            rows.append(tuple(row))
            if len(rows) >= 500:  # safety cap for very large CSVs
                break

    if not rows:
        raise Exception('CSV file is empty.')

    peek = rows[:20]
    return peek, iter(rows[20:]), (lambda: None)


def _parse_csv(path):
    peek, remaining_iter, cleanup = _peek_csv(path)
    try:
        return _run_mapping_and_stream(peek, remaining_iter)
    finally:
        cleanup()


def _peek_xls(path):
    """Peek a legacy .xls file using xlrd."""
    try:
        import xlrd
    except ImportError:
        raise Exception(
            '.xls files require the xlrd library. '
            'Install it with: pip install xlrd'
        )
    try:
        wb = xlrd.open_workbook(path)
    except Exception as e:
        raise Exception(f'Could not open .xls file: {e}')

    # Pick the sheet with the most non-empty cells (same heuristic as _find_data_sheet)
    best_ws    = None
    best_peek  = []
    best_score = -1
    for i in range(wb.nsheets):
        ws   = wb.sheet_by_index(i)
        peek = []
        for r in range(min(20, ws.nrows)):
            peek.append(tuple(ws.cell_value(r, c) for c in range(ws.ncols)))
        score = sum(1 for row in peek for c in row if c is not None and str(c).strip())
        if score > best_score:
            best_score, best_ws, best_peek = score, ws, peek

    if not best_peek:
        raise Exception('XLS file is empty.')

    def remaining_rows():
        for r in range(len(best_peek), best_ws.nrows):
            yield tuple(best_ws.cell_value(r, c) for c in range(best_ws.ncols))

    return best_peek, remaining_rows(), (lambda: None)


def _parse_xls(path):
    peek, remaining_iter, cleanup = _peek_xls(path)
    try:
        return _run_mapping_and_stream(peek, remaining_iter)
    finally:
        cleanup()


def _peek_xlsb(path):
    """Peek a binary .xlsb file using pyxlsb."""
    try:
        from pyxlsb import open_workbook
    except ImportError:
        raise Exception(
            '.xlsb files require the pyxlsb library. '
            'Install it with: pip install pyxlsb'
        )
    try:
        wb = open_workbook(path)
    except Exception as e:
        raise Exception(f'Could not open .xlsb file: {e}')

    best_sheet_name = None
    best_peek       = []
    best_score      = -1

    for name in wb.sheets:
        with wb.get_sheet(name) as ws:
            peek = []
            for row in ws.rows():
                peek.append(tuple(c.v for c in row))
                if len(peek) >= 20:
                    break
        score = sum(1 for row in peek for c in row if c is not None and str(c).strip())
        if score > best_score:
            best_score, best_sheet_name, best_peek = score, name, peek

    wb.close()

    if not best_peek:
        raise Exception('XLSB file is empty.')

    def remaining_rows():
        with open_workbook(path) as wb2:
            with wb2.get_sheet(best_sheet_name) as ws:
                skipped = 0
                for row in ws.rows():
                    if skipped < len(best_peek):
                        skipped += 1
                        continue
                    yield tuple(c.v for c in row)

    return best_peek, remaining_rows(), (lambda: None)


def _parse_xlsb(path):
    peek, remaining_iter, cleanup = _peek_xlsb(path)
    try:
        return _run_mapping_and_stream(peek, remaining_iter)
    finally:
        cleanup()


def _peek_ods(path):
    """Peek an ODS (OpenDocument Spreadsheet) file using odfpy."""
    try:
        from odf.opendocument import load as odf_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
    except ImportError:
        raise Exception(
            '.ods files require the odfpy library. '
            'Install it with: pip install odfpy'
        )

    try:
        doc = odf_load(path)
    except Exception as e:
        raise Exception(f'Could not open .ods file: {e}')

    sheets = doc.spreadsheet.getElementsByType(Table)
    if not sheets:
        raise Exception('ODS file contains no sheets.')

    def _cell_text(cell):
        """Recursively collect plain text from an ODS cell's paragraph nodes."""
        parts = []
        def _walk(node):
            if hasattr(node, 'data'):
                parts.append(node.data)
            for child in getattr(node, 'childNodes', []):
                _walk(child)
        for p in cell.getElementsByType(P):
            _walk(p)
        return ''.join(parts).strip()

    def _read_rows(sheet, max_rows=2000):
        """
        Read all rows from an ODS sheet into a list of tuples.
        Handles table:number-columns-repeated and table:number-rows-repeated.
        Breaks early on a long run of trailing empty rows (ODS padding artifact).
        """
        rows_out = []
        for tr in sheet.getElementsByType(TableRow):
            row_repeat = int(tr.getAttribute('numberrowsrepeated') or 1)
            cells = []
            for tc in tr.getElementsByType(TableCell):
                col_repeat = int(tc.getAttribute('numbercolumnsrepeated') or 1)
                # Very large col_repeat on an empty cell = ODS trailing-column padding; collapse to 1
                val = _cell_text(tc)
                if not val and col_repeat > 64:
                    col_repeat = 1
                cells.extend([val] * col_repeat)
            # Strip trailing empty cells
            while cells and not cells[-1]:
                cells.pop()
            row = tuple(cells) if cells else ()
            # Very large row_repeat on an empty row = ODS trailing-row padding; stop here
            if not row and row_repeat > 10:
                break
            for _ in range(min(row_repeat, max_rows - len(rows_out))):
                rows_out.append(row)
                if len(rows_out) >= max_rows:
                    return rows_out
        return rows_out

    # Pick the sheet with the most non-empty cells (same heuristic as _find_data_sheet)
    best_sheet = None
    best_score = -1
    for sheet in sheets:
        peek = _read_rows(sheet, max_rows=20)
        score = sum(1 for row in peek for c in row if c)
        if score > best_score:
            best_score = score
            best_sheet = sheet

    if best_sheet is None:
        raise Exception('ODS file is empty.')

    all_rows = _read_rows(best_sheet)
    if not all_rows:
        raise Exception('ODS file is empty.')

    peek = all_rows[:20]
    return peek, iter(all_rows[20:]), (lambda: None)


def _parse_ods(path):
    peek, remaining_iter, cleanup = _peek_ods(path)
    try:
        return _run_mapping_and_stream(peek, remaining_iter)
    finally:
        cleanup()


def _peek_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws, peek = _find_data_sheet(wb)

    if not peek:
        wb.close()
        raise Exception('Excel file is empty.')

    # Fresh iterator for streaming; skip the rows already consumed by _find_data_sheet
    row_iter = ws.iter_rows(values_only=True)
    for _ in range(len(peek)):
        try:
            next(row_iter)
        except StopIteration:
            break

    return peek, row_iter, wb.close


def _parse_xlsx(path):
    peek, remaining_iter, cleanup = _peek_xlsx(path)
    try:
        return _run_mapping_and_stream(peek, remaining_iter)
    finally:
        cleanup()


# ── shared AI mapping + row streaming (used by all format loaders) ───────────

def _looks_like_url(s):
    """
    True only for strings that actually parse as an http(s) URL — a real scheme
    and a real network location. Used to detect candidate document links
    regardless of which column they're in (client Excels don't use a
    consistent column name, and — confirmed in production data — the SAME
    column can hold a real URL for one check type and free text like a postal
    address or Aadhaar excerpt for another). Never treat unvalidated text as a
    URL to fetch — that's what caused "Invalid URL: no scheme supplied" errors
    when address/Aadhaar text was previously passed straight to the downloader.
    """
    s = (s or '').strip()
    if not s:
        return False
    try:
        parsed = urllib.parse.urlparse(s)
    except Exception:
        return False
    return parsed.scheme in ('http', 'https') and bool(parsed.netloc)


def _get_mapping(peek):
    """Call the AI column mapper for a single file's peek rows."""
    return _llm_map_columns(peek)


def _run_mapping_and_stream(peek, remaining_iter):
    """
    Given peek rows (list of tuples) and an iterator for the remaining rows,
    call the AI column mapper and stream all data rows into records.
    Shared by _parse_xlsx, _parse_xls, _parse_xlsb, _parse_ods, and _parse_csv.
    """
    mapping = _get_mapping(peek)
    return _stream_rows(peek, remaining_iter, mapping)


def _stream_rows(peek, remaining_iter, mapping, require_identity=True):
    """
    Given peek rows, a remaining-rows iterator, and an already-computed mapping
    ({'header_row_index': ..., 'column_map': ...}), stream all data rows into
    records. Split out from _run_mapping_and_stream so parse_excel_multi() can
    reuse this after a single combined AI call covers multiple files at once.

    require_identity=False (used only by parse_excel_multi) skips the "this file
    has no ARS/Name/Check ID at all" guard — a legitimate supplementary sheet in a
    multi-file batch (e.g. only a join key + a status column) has no identity of
    its own by design; its rows are meant to be resolved later via cross-file
    matching in _cross_link_rows, not rejected before that gets a chance to run.
    """
    header_row_index = mapping.get('header_row_index', 0)
    column_map       = mapping.get('column_map', {})

    if not column_map:
        raise Exception(
            'AI could not identify any columns in this file. '
            'The file may be empty or have an unrecognisable structure.'
        )

    header_row    = peek[header_row_index]
    found_headers = [str(h) for h in header_row if h is not None]

    # Strip whitespace from column_map keys so trailing spaces in headers never cause a miss
    normalized_map = {k.strip(): v for k, v in column_map.items()}

    # Build index: standard_field → list of column positions
    # List because candidate_name can span multiple columns (First Name + Last Name)
    col_index = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        standard_field = normalized_map.get(str(cell).strip())
        if standard_field and standard_field != 'unknown':
            col_index.setdefault(standard_field, []).append(i)

    field_alias = {
        'ars_number':     'ars',
        'check_id':       'check_id',
        'check_type':     'check_type',
        'candidate_name': 'name',
        'emp_id':         'emp_id',
    }
    normalised_index = {field_alias.get(k, k): v for k, v in col_index.items()}

    if (
        require_identity
        and 'ars' not in normalised_index
        and 'name' not in normalised_index
        and 'check_id' not in normalised_index
    ):
        raise Exception(
            f'AI mapped the columns but could not find ARS Number, Candidate Name, or Check ID. '
            f'Columns found: {found_headers}. AI mapping: {column_map}'
        )

    # Full ordered header list — used to capture ALL column values in raw_fields
    full_headers = [
        str(h).strip() if h is not None else f'col_{i}'
        for i, h in enumerate(header_row)
    ]

    def get(row, field):
        indices = normalised_index.get(field, [])
        if not indices:
            return ''
        if field == 'name':
            parts = [
                str(row[i]).strip()
                for i in indices
                if i < len(row) and row[i] is not None and str(row[i]).strip()
            ]
            return ' '.join(parts)
        idx = indices[0]
        if idx >= len(row) or row[idx] is None:
            return ''
        return str(row[idx]).strip()

    records              = []
    _MAX_CONSECUTIVE_EMPTY = 5
    consecutive_empty    = 0

    def _feed(row):
        nonlocal consecutive_empty
        if all(cell is None or str(cell).strip() == '' for cell in row):
            consecutive_empty += 1
            return consecutive_empty < _MAX_CONSECUTIVE_EMPTY
        consecutive_empty = 0

        # Capture every column value keyed by original header — nothing dropped
        raw_fields = {}
        for i, header in enumerate(full_headers):
            if i < len(row) and row[i] is not None and str(row[i]).strip():
                raw_fields[header] = str(row[i]).strip()

        # Document link detection is column-name-agnostic and deterministic —
        # ANY cell in this row that actually parses as an http(s) URL counts,
        # regardless of which column it's in. Confirmed in production data that
        # a single column name (e.g. "Link") can hold a real URL for one check
        # type and unrelated free text (address, Aadhaar excerpt) for another,
        # so this checks every cell's actual value rather than trusting one
        # column to consistently mean "document link" across all rows.
        document_links = [v for v in raw_fields.values() if _looks_like_url(v)]

        record = {
            'ars':             get(row, 'ars'),
            'check_id':        get(row, 'check_id'),
            'check_type':      get(row, 'check_type'),
            'name':            get(row, 'name'),
            'emp_id':          get(row, 'emp_id'),
            'document_links':  document_links,
            'raw_fields':      raw_fields,
        }
        if record['ars']:
            record['folder_key']      = record['ars']
            record['folder_key_type'] = 'ars'
        elif record['check_id']:
            record['folder_key']      = record['check_id']
            record['folder_key_type'] = 'check_id'
        elif record['emp_id']:
            record['folder_key']      = record['emp_id']
            record['folder_key_type'] = 'emp_id'
        elif record['name']:
            record['folder_key']      = record['name']
            record['folder_key_type'] = 'name'
        else:
            record['folder_key']      = ''
            record['folder_key_type'] = 'none'
            record['_flag']           = 'no_identifier'
        records.append(record)
        return True

    # Process the peek rows that follow the header
    _stopped = False
    for row in peek[header_row_index + 1:]:
        if not _feed(row):
            _stopped = True
            break

    # Stream remaining rows (skip if the peek already hit the stop condition)
    if not _stopped:
        for row in remaining_iter:
            if not _feed(row):
                break

    return records
